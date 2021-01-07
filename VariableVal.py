import openpyxl
import datetime
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains

def locate_by_name(web_driver, name):
    """Clicks on something by name."""
    WebDriverWait(web_driver, 10).until(
            EC.presence_of_element_located((By.NAME, name))).click()
    # Returns nothing

def get_by_name(web_driver, name):
    """Returns something by name."""
    return WebDriverWait(web_driver, 10).until(
            EC.presence_of_element_located((By.NAME, name)))

def locate_by_id(web_driver, id):
    """Clicks on something by id."""
    WebDriverWait(web_driver, 10).until(
            EC.presence_of_element_located((By.ID, id))).click()
    # Returns nothing

def get_by_id(web_driver, id):
    """Returns something by id."""
    return WebDriverWait(web_driver, 10).until(
            EC.presence_of_element_located((By.ID, id)))

def locate_by_class(web_driver, class_name):
    """Clicks on something by class name."""
    WebDriverWait(web_driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, class_name))).click()
    # Returns nothing

def get_by_class(web_driver, class_name):
    """Returns something by class name."""
    return WebDriverWait(web_driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, class_name)))

def select_drop(web_driver, id, value):
    """Chooses an option by value in a dropdown box by id."""
    select = Select(get_by_id(web_driver, id))
    select.select_by_value(value)
    # Returns nothing

def standard(date):
    """Make sure that the inputted dates are valid."""
    temp_date = date.split('/')  # Failure to perform the split will result in an exception
    for go in range(3):  # Convert all numbers to ints
        temp_date[go] = int(temp_date[go])

    return datetime.date(temp_date[2], temp_date[0], temp_date[1])

def extract(driver, time1, span):
    results = []  # Will hold all the totals in the range

    select_drop(driver, "fltGroupBy", "BuildingInvType")
    # Click to unselect these
    driver.find_element_by_xpath("//input[@value='CConsign']").click()
    driver.find_element_by_xpath("//input[@value='Supplies']").click()
    driver.find_element_by_xpath("//input[@value='NonStd']").click()
    # Fill out the time boxes with end of workday
    box = get_by_id(driver, "fltTime_TME")
    box.clear()
    box.send_keys("5:00 PM")
    box = get_by_id(driver, "fltCostTime_TME")
    box.clear()
    box.send_keys("5:00 PM")
    for index in range(span):  # For each day in the range
        temp = str(time1.month) + "/" + str(time1.day) + "/" + str(time1.year)
        # For the inventory date
        box = get_by_id(driver, "fltDate_DTE_RQD")
        box.clear()
        box.send_keys(temp)
        # For the cost date
        box = get_by_id(driver, "fltCostDate_DTE_RQD")
        box.clear()
        box.send_keys(temp)

        locate_by_id(driver, "SubmitLink")
        time.sleep(5)
        try:  # Make sure results are present
            get_by_class(driver, "StandardGrid")
        except:  # If not, they probably need a built cache. So "refresh" the page
            locate_by_id(driver, "SubmitLink")

        try:
            default = driver.find_elements_by_class_name("DefaultColor")
            alt = driver.find_elements_by_class_name("AltColor")
            default = default[-1]
            alt = alt[-1]

            if "Total" in default.text:  # If the last row is default color
                total = default.text.split()
            else:  # If the last row is alt color
                total = alt.text.split()

            total.pop(3)  # Remove "Vendor Managed" column
            total.pop(0)  # Remove "Building Code" column
            #for index in range(4):  # For each number in the row
            #    total[index] = float(total[index].replace(",", "_"))
            total.insert(0, temp)

            results.append(total)
        except:
            pass

        time1 += datetime.timedelta(days=1)  # Increment the day

    return results

def Plex(time1, span):
    try:
        # Getting into Plex
        driver = webdriver.Chrome("chromedriver.exe")
        driver.get("https://www.plexonline.com/modules/systemadministration/login/index.aspx?")
        driver.find_element_by_name("txtUserID").send_keys("w.Andre.Le")
        driver.find_element_by_name("txtPassword").send_keys("OokyOoki2")
        driver.find_element_by_name("txtCompanyCode").send_keys("wanco")
        locate_by_id(driver, "btnLogin")
        # Since logging in through selenium opens up a new window instead of changing the current login screen,
        # go to the new screen
        driver.switch_to.window(driver.window_handles[1])
        action = ActionChains(driver)
        action.key_down(Keys.CONTROL).send_keys('M').key_up(Keys.CONTROL).perform()
        action = 404
        time.sleep(1)
        action = ActionChains(driver)
        action.send_keys("Inventory Valuation at Standard Cost").send_keys(Keys.RETURN).perform()
        action = 404
        time.sleep(1)
        action = ActionChains(driver)
        action.send_keys(Keys.DOWN).send_keys(Keys.RETURN).perform()
        time.sleep(1)
        info = extract(driver, time1, span)
    except Exception as e:
        print("An error was encountered:")
        print(e)
    finally:  # Whether the operation was erroneous or successful, quit the driver
        driver.quit()
    
    try:
        return info
    except:
        print("Data missing due to program bug")
        exit()

def main():
    valid = False
    while not valid:
        try:
            date1 = input("Enter in the first date(MM/DD/YYYY): ")
            date1 = standard(date1)
            date2 = input("Enter in the second date(MM/DD/YYYY): ")
            date2 = standard(date2)
            valid = True
        except:
            print("At least one of the dates entered didn't have a valid format or is impossible")

    if date1 > date2:  # If date1 is more recent
        days = (date1 - date2).days + 1
        date = date2
        later_date = date1
    elif date1 < date2:  # If date2 is more recent
        days = (date2 - date1).days + 1
        date = date1
        later_date = date2
    else:  # Both dates are the same
        days = 1
        date = date1
        later_date = date1
    # The "extra" 1 in days is because the first day is included

    data = Plex(date, days)  # Get the data
    wb_obj = openpyxl.Workbook()
    sheet_obj = wb_obj.active
    headers = ["Date", "Finished Goods", "Raw Material", "WIP", "Total"]
    for num, head in enumerate(headers):  # Put in the headers
        sheet_obj.cell(row=1, column=num+1).value = head
    for num, day in enumerate(data):  # Put in the data
        sheet_obj.cell(row=num+2, column=1).value = day[0]
        sheet_obj.cell(row=num+2, column=2).value = day[1]
        sheet_obj.cell(row=num+2, column=3).value = day[2]
        sheet_obj.cell(row=num+2, column=4).value = day[3]
        sheet_obj.cell(row=num+2, column=5).value = day[4]
    title = "Inventory Valuation(" + str(date) + "_to_" + str(later_date) + ").xlsx"
    wb_obj.save(title)

main()